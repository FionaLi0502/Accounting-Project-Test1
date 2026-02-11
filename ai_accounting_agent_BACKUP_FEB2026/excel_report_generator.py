"""
Excel Report Generator Module
Creates professional Excel reports with formulas and formatting
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart
import os
from datetime import datetime

class ExcelReportGenerator:
    """Generate professional Excel reports with industry-standard formatting"""
    
    # Color standards
    COLOR_BLUE = '0000FF'  # Inputs
    COLOR_BLACK = '000000'  # Formulas
    COLOR_GREEN = '008000'  # Internal links
    COLOR_YELLOW = 'FFFF00'  # Assumptions/highlights
    COLOR_HEADER = '4472C4'  # Header background
    COLOR_ALT_ROW = 'F2F2F2'  # Alternating row
    
    def __init__(self):
        """Initialize report generator"""
        self.wb = None
        self.output_dir = "outputs/excel"
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generate_report(self, data: dict) -> str:
        """
        Generate comprehensive Excel report
        
        Args:
            data: Dictionary containing all report data
        
        Returns:
            Path to generated Excel file
        """
        
        self.wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Create worksheets
        self._create_summary_sheet(data)
        self._create_validation_sheet(data)
        self._create_pl_sheet(data)
        self._create_analysis_sheet(data)
        self._create_data_sheet(data)
        
        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        period = data.get('period', 'Report').replace(' ', '_')
        filename = f"Financial_Analysis_{period}_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, filename)
        
        self.wb.save(output_path)
        
        return output_path
    
    def _create_summary_sheet(self, data: dict):
        """Create executive summary sheet"""
        ws = self.wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = data.get('title', 'Financial Analysis Report')
        ws['A1'].font = Font(size=18, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Period: {data.get('period', 'N/A')}"
        ws['A2'].font = Font(size=12, italic=True)
        ws.merge_cells('A2:F2')
        
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A3'].font = Font(size=10, color='666666')
        
        # Key Metrics
        row = 5
        ws[f'A{row}'] = "KEY METRICS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        data_summary = data.get('data_summary', {})
        metrics = [
            ("Total Transactions", data_summary.get('total_transactions', 0)),
            ("Date Range", data_summary.get('date_range', 'N/A')),
            ("Unique Accounts", data_summary.get('accounts', 0)),
            ("Departments", data_summary.get('departments', 0)),
        ]
        
        for metric_name, metric_value in metrics:
            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = metric_value
            ws[f'B{row}'].font = Font(color=self.COLOR_BLUE)
            row += 1
        
        # Data Quality Status
        row += 2
        ws[f'A{row}'] = "DATA QUALITY STATUS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        validation_summary = data.get('validation_summary', {})
        total_issues = sum(len(issues) for issues in validation_summary.values())
        
        ws[f'A{row}'] = "Total Issues Detected"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_issues
        
        if total_issues == 0:
            ws[f'B{row}'].font = Font(color='008000', bold=True)  # Green
            ws[f'C{row}'] = "✓ PASSED"
            ws[f'C{row}'].font = Font(color='008000', bold=True)
        else:
            ws[f'B{row}'].font = Font(color='FF0000', bold=True)  # Red
            ws[f'C{row}'] = "⚠ REVIEW REQUIRED"
            ws[f'C{row}'].font = Font(color='FF6600', bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
    
    def _create_validation_sheet(self, data: dict):
        """Create data validation details sheet"""
        ws = self.wb.create_sheet("Data Validation")
        
        # Header
        ws['A1'] = "DATA VALIDATION REPORT"
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type='solid')
        ws.merge_cells('A1:C1')
        
        # Column headers
        headers = ['Check Category', 'Issue Count', 'Details']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Validation results
        validation_summary = data.get('validation_summary', {})
        row = 4
        
        for category, issues in validation_summary.items():
            ws[f'A{row}'] = category
            ws[f'A{row}'].font = Font(bold=True)
            
            ws[f'B{row}'] = len(issues)
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            
            if len(issues) == 0:
                ws[f'C{row}'] = "✓ No issues detected"
                ws[f'C{row}'].font = Font(color='008000')
            else:
                # List issues
                ws[f'C{row}'] = issues[0] if issues else ""
                for i, issue in enumerate(issues[1:3], start=1):  # Show max 3 issues
                    ws[f'C{row+i}'] = issue
                    ws.merge_cells(f'A{row}:A{row+i}')
                    ws.merge_cells(f'B{row}:B{row+i}')
                row += len(issues[1:3])
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 60
    
    def _create_pl_sheet(self, data: dict):
        """Create P&L statement sheet with formulas"""
        ws = self.wb.create_sheet("P&L Statement")
        
        # Header
        ws['A1'] = "PROFIT & LOSS STATEMENT"
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type='solid')
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Period: {data.get('period', 'N/A')}"
        ws['A2'].font = Font(italic=True)
        
        # Column headers
        headers = ['Line Item', 'Amount ($)', '% of Revenue', 'Notes']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type='solid')
        
        # P&L data
        pl_statement = data.get('pl_statement')
        if pl_statement is not None and not pl_statement.empty:
            row = 5
            revenue_row = None
            
            for idx, pl_row in pl_statement.iterrows():
                line_item = pl_row['Line Item']
                amount = pl_row.get('Amount')
                pct = pl_row.get('% of Revenue')
                
                ws[f'A{row}'] = line_item
                
                # Format based on line type
                if 'Total' in line_item or line_item in ['Gross Profit', 'Operating Income', 'Net Income']:
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'].font = Font(bold=True)
                    
                    # Add borders
                    thin_border = Border(top=Side(style='thin'))
                    ws[f'A{row}'].border = thin_border
                    ws[f'B{row}'].border = thin_border
                    ws[f'C{row}'].border = thin_border
                
                # Save revenue row for percentage calculations
                if line_item == 'Total Revenue':
                    revenue_row = row
                
                # Amount
                if amount is not None and pd.notna(amount):
                    ws[f'B{row}'] = amount
                    ws[f'B{row}'].number_format = '$#,##0;($#,##0);-'
                    
                    # Use formula for percentages instead of hardcoded values
                    if revenue_row and amount != 0:
                        ws[f'C{row}'] = f'=B{row}/B${revenue_row}*100'
                        ws[f'C{row}'].number_format = '0.0'
                        ws[f'C{row}'].font = Font(color=self.COLOR_BLACK)  # Black for formulas
                
                row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
    
    def _create_analysis_sheet(self, data: dict):
        """Create AI analysis sheet"""
        ws = self.wb.create_sheet("AI Analysis")
        
        # Header
        ws['A1'] = "AI-GENERATED INSIGHTS"
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type='solid')
        ws.merge_cells('A1:F1')
        
        # AI insights
        ai_insights = data.get('ai_insights', 'No analysis available')
        
        # Split into lines and format
        lines = ai_insights.split('\n')
        row = 3
        
        for line in lines:
            if line.strip():
                # Check if it's a header (starts with # or is all caps)
                if line.startswith('#') or (line.isupper() and len(line) < 50):
                    ws[f'A{row}'] = line.replace('#', '').strip()
                    ws[f'A{row}'].font = Font(size=12, bold=True, color=self.COLOR_HEADER)
                else:
                    ws[f'A{row}'] = line.strip()
                    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                
                row += 1
        
        # Set column width
        ws.column_dimensions['A'].width = 100
    
    def _create_data_sheet(self, data: dict):
        """Create raw data sheet"""
        ws = self.wb.create_sheet("Raw Data")
        
        # This would contain the raw GL data
        # For now, just add a placeholder
        ws['A1'] = "Raw GL data would be included here"
        ws['A1'].font = Font(italic=True, color='666666')
