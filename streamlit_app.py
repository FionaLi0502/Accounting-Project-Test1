"""Three Statements Automation - Streamlit App (patched)
Focus:
- TB + GL must be treated as a set for generation.
- Validation section is always visible and detailed.
- Auto-fix is user-controlled (accept/reject) instead of silently applied.
- Demo controls:
  1) Download Sample Financial Model (SAMPLE_DEMO template)
  2) Download Sample Dataset (current loaded TB+GL as a zip)
  3) Load Random Sample Set (TB+GL together, matched years)
"""

import io
import zipfile
from datetime import datetime

import pandas as pd
import openpyxl
import numpy as np
import streamlit as st

from validation import (
    validate_trial_balance,
    validate_gl_activity,
    apply_auto_fixes,
    validate_year0_opening_snapshot,
    add_year0_snapshot,
)
from mapping import map_accounts, TEMPLATE_LABEL_MAPPING
from excel_writer import (
    calculate_3statements_from_tb_gl,
    calculate_financial_statements,
    write_financial_data_to_template,
    compute_reconciliation_checks,
)
from sample_data import get_sample_data_path, get_template_path
from pdf_export import generate_pdf_report
from ai_summary import generate_ai_summary
import os


# ----------------------------
# Config / defaults
# ----------------------------
st.set_page_config(page_title="AI Accounting Agent", layout="wide")

AUTO_FIX_OPTIONS = [
    ("remove_missing_dates", "Remove rows with missing TxnDate"),
    ("remove_future_dates", "Remove rows with future TxnDate"),
    ("map_unclassified", "Map missing AccountNumber → 9999"),
    ("fix_account_numbers", "Convert negative AccountNumber → positive"),
    ("remove_duplicates", "Remove fully duplicate rows (safe)"),
]

UNIT_SCALE_OPTIONS = {
    "USD (no scaling)": 1,
    "USD thousands (÷ 1000)": 1000,
}

# Session state init
for k, default in {
    "tb_df": None,
    "gl_df": None,
    "tb_name": None,
    "gl_name": None,
    "tb_changes": [],
    "gl_changes": [],
    "validation_tb": [],
    "validation_gl": [],
    "selected_fixes": [],
    "unit_scale": 1000,
    "strict_mode": True,
    "template_type": "zero",  # 'zero' or 'demo'
    "dataset_source": None,  # 'random' or 'upload'
    "preview_sections": {},
    "last_excel_bytes": None,
}.items():
    if k not in st.session_state:
        st.session_state[k] = default


def _issues_to_table(issues):
    """Compact table for issue list."""
    rows = []
    for it in issues:
        rows.append({
            "severity": it.get("severity", ""),
            "category": it.get("category", ""),
            "issue": it.get("issue", ""),
            "impact": it.get("impact", ""),
            "suggestion": it.get("suggestion", ""),
            "auto_fix": it.get("auto_fix", ""),
        })
    return pd.DataFrame(rows)


def _count_by_severity(issues):
    counts = {"Critical": 0, "Warning": 0, "Info": 0}
    for it in issues:
        sev = it.get("severity", "Info")
        if sev not in counts:
            counts[sev] = 0
        counts[sev] += 1
    return counts


def _find_row_exact(ws, label: str, start_row: int = 1, end_row: int = 200):
    target = str(label).strip().lower()
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        if str(v).strip().lower() == target:
            return r
    return None


def _extract_labels(ws, start_row: int, end_row: int) -> list[str]:
    labels = []
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            labels.append("")
        else:
            labels.append(str(v).strip())
    return labels


def _compute_template_preview_sections(financial_data: dict, template_path: str, unit_scale: int = 1):
    """
    Build template-matching preview tables (Income Statement / Balance Sheet / Cash Flow)
    using the template's label order as the source of truth.

    Returns:
        dict[str, pd.DataFrame] sections
    """
    if not financial_data:
        return {}

    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb["Blank 3 Statement Model"] if "Blank 3 Statement Model" in wb.sheetnames else wb[wb.sheetnames[0]]

    years_all = sorted(financial_data.keys())
    stmt_years = years_all[1:] if len(years_all) > 1 else years_all  # hide Year0 in preview

    # Helper: fetch value by internal key
    def v(y: int, key: str) -> float:
        try:
            return float(financial_data.get(y, {}).get(key, 0.0) or 0.0) / float(unit_scale)
        except Exception:
            return 0.0

    # Derived calculations (match template logic)
    def gross_profit(y): return v(y, "revenue") - v(y, "cogs")
    def total_opex(y): 
        return v(y,"distribution_expenses")+v(y,"marketing_admin")+v(y,"research_dev")+v(y,"depreciation_expense")
    def ebit(y): return gross_profit(y) - total_opex(y)
    def ebt(y): return ebit(y) - v(y,"interest_expense")
    def net_income(y): 
        # prefer calculated if present, else compute
        ni = v(y,"net_income")
        return ni if abs(ni) > 1e-9 else (ebt(y) - v(y,"tax_expense"))

    def total_current_assets(y):
        return v(y,"cash")+v(y,"accounts_receivable")+v(y,"inventory")+v(y,"prepaid_expenses")+v(y,"other_current_assets")
    def net_ppe(y):
        return v(y,"ppe_gross") - v(y,"accumulated_depreciation")
    def total_assets(y):
        return total_current_assets(y) + net_ppe(y)

    def total_current_liab(y):
        return (v(y,"accounts_payable")+v(y,"accrued_payroll")+v(y,"deferred_revenue")+
                v(y,"interest_payable")+v(y,"other_current_liabilities")+v(y,"income_taxes_payable"))
    def total_equity(y):
        return v(y,"common_stock") + v(y,"retained_earnings")
    def total_le(y):
        return total_current_liab(y) + v(y,"long_term_debt") + total_equity(y)

    # Cash Flow
    def cfo(y):
        return (net_income(y) + v(y,"depreciation_expense") +
                v(y,"delta_ar")+v(y,"delta_inventory")+v(y,"delta_prepaid")+v(y,"delta_other_current_assets")+
                v(y,"delta_ap")+v(y,"delta_accrued_payroll")+v(y,"delta_deferred_revenue")+v(y,"delta_interest_payable")+
                v(y,"delta_other_current_liabilities")+v(y,"delta_income_taxes_payable"))
    def cfi(y): return v(y,"capex")
    def cff(y): return v(y,"stock_issuance") + (-v(y,"dividends")) + v(y,"delta_debt")
    def net_change_cash(y): return cfo(y) + cfi(y) + cff(y)
    def begin_cash(y):
        bc = v(y, "beginning_cash")
        if bc != 0:
            return bc
        # Derive beginning cash from prior-year cash (Year0 for first statement year)
        if y in stmt_years:
            idx = stmt_years.index(y)
            prev = year0 if idx == 0 else stmt_years[idx - 1]
            return v(prev, "cash")
        return 0.0
    def end_cash(y):
        # prefer key if exists, else compute
        ec = v(y,"cash")
        return ec if abs(ec) > 1e-9 else (begin_cash(y) + net_change_cash(y))

    # Checks (match your Row3/Row81 intent)
    checks = compute_reconciliation_checks(financial_data)
    def bs_check(y): return float(checks.get(y, {}).get("balance_sheet_check", 0.0) or 0.0) / float(unit_scale)
    def cf_check(y): return float(checks.get(y, {}).get("cashflow_check", 0.0) or 0.0) / float(unit_scale)

    # Map template label -> internal key (reverse mapping)
    label_to_key = TEMPLATE_LABEL_MAPPING

    # Build a section DataFrame from template row range
    def build_df(start_row: int, end_row: int):
        labels = _extract_labels(ws, start_row, end_row)
        data_rows = []
        for label in labels:
            row_dict = {}
            norm_label = label.strip()

            # Fetch base value or derived value
            key = label_to_key.get(norm_label)
            if key:
                # Base mapped key
                for y in stmt_years:
                    row_dict[f"FY{y}"] = v(y, key)
            else:
                # Derived or special row
                if norm_label == "Gross Profit":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = gross_profit(y)
                elif norm_label == "Total Operating Expenses":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = total_opex(y)
                elif norm_label == "EBIT (Operating Profit)":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = ebit(y)
                elif norm_label == "Income Before Taxes":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = ebt(y)
                elif norm_label == "Net Income":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = net_income(y)

                elif norm_label == "Total Current Assets":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = total_current_assets(y)
                elif norm_label == "Property, Plant & Equipment - Net":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = net_ppe(y)
                elif norm_label == "TOTAL ASSETS":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = total_assets(y)
                elif norm_label == "Total Current Liabilities":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = total_current_liab(y)
                elif norm_label == "Total Shareholders' Equity":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = total_equity(y)
                elif norm_label == "TOTAL LIABILITIES AND EQUITY":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = total_le(y)

                elif norm_label == "Cash from Operating Activities":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = cfo(y)
                elif norm_label == "Cash from Investing Activities":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = cfi(y)
                elif norm_label == "Cash from Financing Activities":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = cff(y)
                elif norm_label == "Net Change in Cash":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = net_change_cash(y)
                elif norm_label == "Cash and Equivalents, Beginning of the Year":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = begin_cash(y)
                elif norm_label == "Cash and Equivalents, End of the Year":
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = end_cash(y)
                else:
                    # Section headers and other unmapped rows → None
                    for y in stmt_years:
                        row_dict[f"FY{y}"] = None

            row_dict[""] = label  # First column = label
            data_rows.append(row_dict)

        return pd.DataFrame(data_rows).set_index("")

    year0 = years_all[0] if years_all else None

    # Locate sections using template row markers
    is_header = _find_row_exact(ws, "Income Statement", start_row=1, end_row=100)
    is_start = _find_row_exact(ws, "Revenues", start_row=is_header or 1, end_row=140)
    is_end = _find_row_exact(ws, "Common Dividends", start_row=is_start or 1, end_row=160)

    bs_header = _find_row_exact(ws, "Balance Sheet", start_row=40, end_row=120)  # should find row 48
    bs_start = _find_row_exact(ws, "ASSETS", start_row=bs_header or 1, end_row=200)
    bs_end = _find_row_exact(ws, "Check", start_row=bs_start or 1, end_row=140)  # row 81 (check line)

    cf_header = _find_row_exact(ws, "Cash Flow Statement", start_row=70, end_row=160)  # row 84
    cf_start = _find_row_exact(ws, "Cash Flows from Operating Activities:", start_row=cf_header or 1, end_row=200)
    cf_end = _find_row_exact(ws, "Cash and Equivalents, End of the Year", start_row=cf_start or 1, end_row=220)

    sections = {}

    # Income Statement (include from Revenues down to Common Dividends)
    if is_start and is_end:
        sections["Income Statement"] = build_df(is_start, is_end)
    # Balance Sheet (include from ASSETS down to TOTAL L+E)
    if bs_start and bs_end:
        # bs_end is 'Check' row; include up to TOTAL L+E row (one above check)
        sections["Balance Sheet"] = build_df(bs_start, bs_end - 2)
    # Cash Flow (include from Operating Activities down to End of Year)
    if cf_start and cf_end:
        sections["Cash Flow Statement"] = build_df(cf_start, cf_end)

    # Checks section (explicit)
    checks_df = pd.DataFrame(
        {
            f"FY{y}": {
                "Balance Sheet Check (A - L + E)": bs_check(y),
                "Cash Tie-out Check": cf_check(y),
            }
            for y in stmt_years
        }
    )
    sections["Checks"] = checks_df

    return sections


def run_validation():
    """Run validations and store results in session state."""
    tb_df = st.session_state["tb_df"]
    gl_df = st.session_state["gl_df"]

    tb_issues = validate_trial_balance(tb_df) if tb_df is not None else []
    gl_issues = validate_gl_activity(gl_df) if gl_df is not None else []

    st.session_state["validation_tb"] = tb_issues
    st.session_state["validation_gl"] = gl_issues

    return tb_issues, gl_issues


def load_random_set():
    """Load matched TB + GL backup sets by choosing one year-range and loading both files."""
    year_ranges = [(2020, 2022), (2021, 2023), (2022, 2024), (2023, 2025), (2024, 2026)]
    import random
    y0, y1 = random.choice(year_ranges)
    tb_file = f"backup_tb_{y0}_{y1}.csv"
    gl_file = f"backup_gl_{y0}_{y1}_with_txnid.csv"

    tb_path = get_sample_data_path(tb_file)
    gl_path = get_sample_data_path(gl_file)

    tb_df = pd.read_csv(tb_path)
    gl_df = pd.read_csv(gl_path)

    st.session_state["tb_df"] = tb_df
    st.session_state["gl_df"] = gl_df
    st.session_state["tb_name"] = tb_file
    st.session_state["gl_name"] = gl_file
    st.session_state["dataset_source"] = "random"
    st.session_state["validation_tb"] = []
    st.session_state["validation_gl"] = []


tb_df = st.session_state["tb_df"]
gl_df = st.session_state["gl_df"]


# ----------------------------
# App Layout
# ----------------------------
st.title("AI Accounting Agent: 3-Statement Financial Model Generator")
st.markdown("Upload Trial Balance and General Ledger data to auto-generate financial statements.")

st.divider()

# ----------------------------
# Demo Controls
# ----------------------------
st.header("Quick Start / Demo")

demo_cols = st.columns(3)
with demo_cols[0]:
    # Download the sample financial model template (SAMPLE_DEMO template)
    template_demo_path = get_template_path("demo")
    with open(template_demo_path, "rb") as f:
        st.download_button(
            "📥 Download Sample Financial Model (Excel)",
            data=f,
            file_name="sample_financial_model.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

with demo_cols[1]:
    # Download current TB + GL data as a .zip
    if tb_df is not None or gl_df is not None:
        bio = io.BytesIO()
        with zipfile.ZipFile(bio, "w") as z:
            if tb_df is not None:
                tb_csv = tb_df.to_csv(index=False)
                z.writestr(st.session_state.get("tb_name", "trial_balance.csv"), tb_csv)
            if gl_df is not None:
                gl_csv = gl_df.to_csv(index=False)
                z.writestr(st.session_state.get("gl_name", "general_ledger.csv"), gl_csv)
        st.download_button(
            "📥 Download Current Dataset (TB+GL)",
            data=bio.getvalue(),
            file_name="dataset.zip",
            mime="application/zip",
        )
    else:
        st.caption("(Load TB/GL first to enable dataset download)")

with demo_cols[2]:
    if st.button("🎲 Load Random Sample Set (TB+GL)"):
        load_random_set()
        run_validation()
        st.success("Loaded a random TB+GL matched set.")

st.divider()

# ----------------------------
# Upload Files
# ----------------------------
st.header("Upload Files")

st.markdown("**Upload Trial Balance and General Ledger files** (both required as a set for generation).")

col_tb, col_gl = st.columns(2)
with col_tb:
    tb_up = st.file_uploader("Trial Balance (CSV)", type="csv", key="tb_uploader")
    if tb_up:
        st.session_state["tb_df"] = pd.read_csv(tb_up)
        st.session_state["tb_name"] = tb_up.name
        st.session_state["dataset_source"] = "upload"
        run_validation()

with col_gl:
    gl_up = st.file_uploader("General Ledger (CSV)", type="csv", key="gl_uploader")
    if gl_up:
        st.session_state["gl_df"] = pd.read_csv(gl_up)
        st.session_state["gl_name"] = gl_up.name
        st.session_state["dataset_source"] = "upload"
        run_validation()

# Show loaded file summary
if tb_df is not None:
    st.caption(f"**TB:** {st.session_state.get('tb_name')} ({len(tb_df)} rows)")
if gl_df is not None:
    st.caption(f"**GL:** {st.session_state.get('gl_name')} ({len(gl_df)} rows)")

st.divider()

# ----------------------------
# Settings
# ----------------------------
st.header("Settings")

col_set1, col_set2, col_set3 = st.columns(3)
with col_set1:
    unit_choice = st.selectbox("Unit Scale", options=list(UNIT_SCALE_OPTIONS.keys()), index=1)
    st.session_state["unit_scale"] = UNIT_SCALE_OPTIONS[unit_choice]
with col_set2:
    strict = st.checkbox("Strict Mode (fail on Critical issues)", value=st.session_state["strict_mode"])
    st.session_state["strict_mode"] = strict
with col_set3:
    tpl = st.radio("Template Type", options=["zero", "demo"], index=0, horizontal=True)
    st.session_state["template_type"] = tpl

st.divider()

# ----------------------------
# Validation Section (always visible and detailed)
# ----------------------------
st.header("Validation")

if tb_df is not None or gl_df is not None:
    tb_issues = st.session_state["validation_tb"]
    gl_issues = st.session_state["validation_gl"]

    # Severity counts
    tb_counts = _count_by_severity(tb_issues)
    gl_counts = _count_by_severity(gl_issues)

    # Show summary metrics
    st.markdown("### Validation Results")
    col_sum1, col_sum2 = st.columns(2)
    with col_sum1:
        st.markdown("**Trial Balance**")
        st.caption(f"🔴 Critical: {tb_counts['Critical']} | 🟡 Warning: {tb_counts['Warning']} | ℹ️ Info: {tb_counts['Info']}")
    with col_sum2:
        st.markdown("**General Ledger**")
        st.caption(f"🔴 Critical: {gl_counts['Critical']} | 🟡 Warning: {gl_counts['Warning']} | ℹ️ Info: {gl_counts['Info']}")

    # Detailed issues table
    if tb_issues or gl_issues:
        with st.expander("View Detailed Validation Issues", expanded=False):
            if tb_issues:
                st.markdown("#### Trial Balance Issues")
                st.dataframe(_issues_to_table(tb_issues), use_container_width=True)
            if gl_issues:
                st.markdown("#### General Ledger Issues")
                st.dataframe(_issues_to_table(gl_issues), use_container_width=True)

    # Auto-fix controls
    st.markdown("### Auto-fix (accept/reject)")
    detected_fixes = sorted({it.get("auto_fix") for it in (tb_issues + gl_issues) if it.get("auto_fix")})
    if not detected_fixes:
        st.caption("No auto-fixes suggested by the current validation results.")
    else:
        fix_labels = {code: label for code, label in AUTO_FIX_OPTIONS}
        default_selected = [f for f in detected_fixes if f != "remove_duplicates"]  # safer default
        selected = st.multiselect(
            "Select fixes to apply",
            options=detected_fixes,
            default=default_selected,
            format_func=lambda x: f"{x} — {fix_labels.get(x, '')}",
        )

        cA, cB = st.columns([1, 1])
        with cA:
            if st.button("Apply selected fixes"):
                if tb_df is not None:
                    fixed, changes = apply_auto_fixes(tb_df, selected_fixes=selected)
                    st.session_state["tb_df"] = fixed
                    st.session_state["tb_changes"] = changes
                if gl_df is not None:
                    fixed, changes = apply_auto_fixes(gl_df, selected_fixes=selected)
                    st.session_state["gl_df"] = fixed
                    st.session_state["gl_changes"] = changes
                run_validation()
                st.success("Applied selected fixes and re-ran validation.")
        with cB:
            if st.button("Re-run validation"):
                run_validation()
                st.info("Re-ran validation.")

        if st.session_state.get("tb_changes"):
            st.caption("TB auto-fixes applied:")
            st.write(st.session_state["tb_changes"])
        if st.session_state.get("gl_changes"):
            st.caption("GL auto-fixes applied:")
            st.write(st.session_state["gl_changes"])


# ----------------------------
# Generation
# ----------------------------
st.divider()
st.header("Generate Outputs")

strict_mode = st.session_state["strict_mode"]

if st.button("Generate 3-Statement Outputs", type="primary"):
    # Require TB + GL as a set for generation
    if tb_df is None or gl_df is None:
        st.error("TB and GL must be loaded as a set to generate the 3-statement output.")
        st.stop()

    # Strict mode gate: fail if any Critical issues remain
    tb_issues, gl_issues = run_validation()
    crit = [it for it in (tb_issues + gl_issues) if it.get("severity") == "Critical"]
    if strict_mode and crit:
        st.error("Strict mode: Critical validation issues must be resolved before generation. See Validation details above.")
        st.stop()

    # Year0 strict requirement
    if strict_mode:
        year0_issues = validate_year0_opening_snapshot(tb_df, statement_years=3)
        if year0_issues:
            # If the dataset came from the built-in random demo/backups, we can create an internal Year0 snapshot
            # (you asked for Year0 to exist for internal use). For user uploads, we do NOT synthesize Year0.
            if st.session_state.get("dataset_source") == "random":
                tb_fixed, msg = add_year0_snapshot(tb_df, statement_years=3)
                st.session_state["tb_df"] = tb_fixed
                tb_df = tb_fixed
                st.info(msg)
                run_validation()
                year0_issues = validate_year0_opening_snapshot(tb_df, statement_years=3)

            if year0_issues:
                st.error("Strict mode: Year0 opening snapshot requirement failed:\n" + "\n".join(year0_issues))
                st.stop()

    # Map accounts
    tb_mapped = map_accounts(tb_df)
    gl_mapped = map_accounts(gl_df)

    # Compute 3-statement data
    financial_data = calculate_3statements_from_tb_gl(tb_mapped, gl_mapped)

    # Write to template
    template_path = get_template_path(st.session_state["template_type"])
    out_bytes = write_financial_data_to_template(
        template_path=template_path,
        financial_data=financial_data,
        unit_scale=float(st.session_state["unit_scale"]),
    )

    # Persist output so Streamlit reruns don't lose it
    st.session_state["last_excel_bytes"] = out_bytes.getvalue()

    # Build a template-matching preview (Income Statement / Balance Sheet / Cash Flow)
    try:
        template_path_for_preview = get_template_path(st.session_state.get("template_type", "zero"))
        st.session_state["preview_sections"] = _compute_template_preview_sections(
            financial_data=financial_data,
            template_path=template_path_for_preview,
            unit_scale=st.session_state.get("unit_scale", 1),
        )
    except Exception:
        st.session_state["preview_sections"] = {}

    st.success("Generated Excel output.")
    st.download_button(
        "Download Excel Output",
        data=out_bytes.getvalue(),
        file_name="3statement_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ========================================
    # 3-Statement Preview (moved here from bottom)
    # ========================================
    if st.session_state.get("preview_sections"):
        st.subheader("3-Statement Output Preview (Template-Matching)")

        sections = st.session_state["preview_sections"]
        tab_names = [k for k in ["Income Statement", "Balance Sheet", "Cash Flow Statement", "Checks"] if k in sections]
        if tab_names:
            tabs = st.tabs(tab_names)
            for tname, tab in zip(tab_names, tabs):
                with tab:
                    df = sections.get(tname)
                    if df is None or (hasattr(df, "empty") and df.empty):
                        st.info("No preview data available for this section.")
                    else:
                        # Clean display: Streamlit shows Python None as 'None' which is misleading for section headers.
                        # For website preview only: show blanks for headers and 0 for unmapped numeric rows (already 0 in df).
                        df_show = df.copy()
                        df_show = df_show.replace({None: np.nan})

                        def _fmt_cell(x):
                            if pd.isna(x):
                                return ""
                            if isinstance(x, (int, float, np.integer, np.floating)):
                                return f"{float(x):,.0f}"
                            return str(x)

                        df_show = df_show.applymap(_fmt_cell)
                        st.dataframe(df_show, use_container_width=True)

    # ========================================
    # AI Summary Generation
    # ========================================
    try:
        # Get API key from Streamlit secrets or environment
        api_key = None
        try:
            api_key = st.secrets.get("ANTHROPIC_API_KEY")
        except:
            api_key = os.environ.get("ANTHROPIC_API_KEY")
        
        # Determine data availability
        has_balance_sheet = (tb_df is not None)
        has_cash_flow = (tb_df is not None and len(financial_data) >= 2)
        
        # Generate summary
        summary_text, used_ai = generate_ai_summary(
            financial_data=financial_data,
            has_balance_sheet=has_balance_sheet,
            has_cash_flow=has_cash_flow,
            api_key=api_key
        )
        
        # Display summary
        st.subheader("📊 AI Financial Summary")
        if used_ai:
            st.success("✅ Generated using Claude AI")
        else:
            st.info("ℹ️ Generated using rule-based analysis (AI unavailable)")
        
        st.markdown(summary_text)
        
    except Exception as e:
        st.warning(f"Could not generate AI summary: {str(e)}")
        summary_text = None
    
    # ========================================
    # PDF Report Generation
    # ========================================
    try:
        pdf_bytes = generate_pdf_report(
            financial_data=financial_data,
            summary_text=summary_text if summary_text else "Summary not available",
            unit_scale=st.session_state.get("unit_scale", 1000)
        )
        
        st.download_button(
            "📄 Download PDF Report",
            data=pdf_bytes,
            file_name="financial_report.pdf",
            mime="application/pdf",
        )
        
    except Exception as e:
        st.warning(f"Could not generate PDF: {str(e)}")


# Persisted output preview / download (survives reruns)
if st.session_state.get("last_excel_bytes"):
    st.download_button(
        "Download Excel Output (last run)",
        data=st.session_state["last_excel_bytes"],
        file_name="3statement_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_last_excel",
    )